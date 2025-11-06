import requests
from openpyxl import load_workbook
from datetime import datetime, timedelta
import logging
logger = logging.getLogger(__name__)

def generate_excel(json_file: dict, company: str):

    try:
        wb = load_workbook(f'./xlsx_files/{company}/template.xlsx')
        today = datetime.today().date().strftime("%Y%m%d")
        destination = f"./xlsx_files/{company}/{today}_uzupelniony.xlsx"

        logger.info("Template file loaded")
    except requests.exceptions.RequestException as e:
        print("❌ Template file not loaded:", e)
        return



    ws = wb["Dane"]

    # Pobierz godziny z wiersza 11 (C11:Z11)
    hour_row = 11
    hour_cols = list(range(3, 27))  # kolumny C (3) do Z (26)
    excel_hours = [ws.cell(row=hour_row, column=col).value for col in hour_cols]
    logger.debug("Godziny odczytane z arkusza: %s", excel_hours)

    # Wypełnij sekcję POBIERANA IRRADIANCJA (B12:Z22)
    logger.info("Uzupełniam sekcję POBIERANA IRRADIANCJA")
    start_row = 12
    for i, (date, hours) in enumerate(json_file.items()):
        row = start_row + i
        ws.cell(row=row, column=2).value = date  # kolumna B z datą
        for col_index, excel_hour in zip(hour_cols, excel_hours):
            json_hour = "00:00" if excel_hour == 24 else f"{int(excel_hour):02d}:00"
            value = hours.get(json_hour, 0)
            ws.cell(row=row, column=col_index).value = value

    # Pobierz irradiancję 100% i netto 100% z wierszy 6 i 7 (C6:Z6 i C7:Z7)
    irradiance_100 = [ws.cell(row=6, column=col).value for col in hour_cols]
    netto_100 = [ws.cell(row=7, column=col).value for col in hour_cols]
    logger.debug("Odczytano irradiancję 100%% i netto 100%%.")

    # Wypełnij sekcję Obliczenia (B23:Z44) – „Netto” i „Do sieci”
    logger.info("Rozpoczynam obliczenia Netto i Do sieci")
    calc_start_row = 23
    for i, (date, hours) in enumerate(json_file.items()):
        netto_row = calc_start_row + i * 2
        do_sieci_row = netto_row + 1
        ws.cell(row=netto_row, column=2).value = "Netto"
        ws.cell(row=netto_row, column=1).value = date
        ws.cell(row=do_sieci_row, column=2).value = "Do sieci"
        ws.cell(row=do_sieci_row, column=1).value = date

        try:
            date_obj = datetime.strptime(date, "%d.%m.%Y")
        except ValueError:
            logger.warning("Nie udało się sparsować daty: %s", date)
            continue

        weekday = date_obj.weekday()
        if weekday < 5:
            needs_row = 3
        elif weekday == 5:
            needs_row = 4
        else:
            needs_row = 5

        for idx, (col_index, excel_hour) in enumerate(zip(hour_cols, excel_hours)):
            json_hour = "00:00" if excel_hour == 24 else f"{int(excel_hour):02d}:00"
            irradiance_value = hours.get(json_hour, 0)
            irr_100 = ws.cell(row=6, column=col_index).value or 1
            net_100 = ws.cell(row=7, column=col_index).value or 0
            netto = (irradiance_value / irr_100) * net_100 if irr_100 else 0
            ws.cell(row=netto_row, column=col_index).value = round(netto, 4)

            needs_value = ws.cell(row=needs_row, column=col_index).value or 0
            do_sieci = max(netto - needs_value, 0)
            ws.cell(row=do_sieci_row, column=col_index).value = round(do_sieci, 4)

    # Dodanie 4 kolejnych dni
    if json_file:
        logger.info("Dodaję 4 kolejne dni na podstawie ostatniego dnia z JSON-a")
        last_date_str = max(json_file, key=lambda d: datetime.strptime(d, "%d.%m.%Y"))
        last_date = datetime.strptime(last_date_str, "%d.%m.%Y")

        last_netto_row = calc_start_row + (len(json_file) - 1) * 2
        last_do_sieci_row = last_netto_row + 1

        last_netto_values = [ws.cell(row=last_netto_row, column=col).value for col in hour_cols]
        last_do_sieci_values = [ws.cell(row=last_do_sieci_row, column=col).value for col in hour_cols]

        last_row = calc_start_row + len(json_file) * 2

        for i in range(1, 5):
            new_date = (last_date + timedelta(days=i)).strftime("%d.%m.%Y")
            netto_row = last_row + (i - 1) * 2
            do_sieci_row = netto_row + 1
            ws.cell(row=netto_row, column=1).value = new_date
            ws.cell(row=netto_row, column=2).value = "Netto"
            ws.cell(row=do_sieci_row, column=1).value = new_date
            ws.cell(row=do_sieci_row, column=2).value = "Do sieci"

            for col_index, (val_netto, val_do_sieci) in zip(hour_cols, zip(last_netto_values, last_do_sieci_values)):
                ws.cell(row=netto_row, column=col_index).value = val_netto
                ws.cell(row=do_sieci_row, column=col_index).value = val_do_sieci

    # Kopiowanie tabeli do arkusza "Grafiki generacji"
    logger.info("Kopiuję tabelę z arkusza 'Dane' do arkusza 'Grafiki generacji'")
    src_ws = wb["Dane"]
    dst_ws = wb["Grafiki generacji"]

    src_start_row, src_end_row = 22, 44
    src_start_col, src_end_col = 1, 26
    dst_start_row = 8
    dst_start_col = 1

    for r_offset, row in enumerate(range(src_start_row, src_end_row + 1)):
        for c_offset, col in enumerate(range(src_start_col, src_end_col + 1)):
            value = src_ws.cell(row=row, column=col).value
            dst_ws.cell(row=dst_start_row + r_offset, column=dst_start_col + c_offset).value = value

    wb.save(destination)
    logger.info("Zapisano wynik do pliku: %s", destination)

    return destination



